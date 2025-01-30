"""Vamos a poder trabajar con excel"""
import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")

print(wb.sheetnames) #trae la lista de nombres de las hojas
print(wb["nombre de la hoja"])

hoja = wb.active #trae la hoja activa
print(hoja)

wb.create_sheet("nombre de la hoja nueva") #para agregar una hoja nueva

hojaNueva = wb["nombre de hoja nueva"]
hojaNueva.title = "Nuevo Titulo"

#podemos ver la cantidad de filas y columnas activas
print(
    hoja.max_row,
    hoja.max_column
)

celda = hoja["A1"]
print(celda.value) #imprime la celda y el valor que contiene

"""
celda2 = hoja.cell(row=2, column=1)
print(celda2.value)

LAS CELDAS FILAS Y COLUMNAS PARTEN DEL INDICE 1 (NO DESDE CERO)

"""
#PARA ITERAR EN PLANILLAS ENTONCES
for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column +1):
        celda = hoja.cell(row=fila, column=columna)
        print(fila, columna, celda.value)

#puede que quiera buscar un fila y una columna en particular

